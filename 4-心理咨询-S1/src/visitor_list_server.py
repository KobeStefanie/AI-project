#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
来访者列表数据服务
提供来访者列表查询和Excel导出功能
端口: 8770
"""

from flask import Flask, jsonify, send_file
from flask_cors import CORS
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

app = Flask(__name__)
CORS(app)

# 项目根目录
PROJECT_ROOT = Path(__file__).parent.parent
OUTPUT_DIR = PROJECT_ROOT / 'output' / '来访者库'
DATA_DIR = PROJECT_ROOT / 'data' / 'cases' / 'processed'


def get_visitor_folders():
    """获取所有来访者文件夹"""
    if not OUTPUT_DIR.exists():
        return []

    visitor_folders = []
    for folder in OUTPUT_DIR.iterdir():
        if folder.is_dir() and folder.name.startswith('V'):
            visitor_folders.append(folder)

    return sorted(visitor_folders, key=lambda x: x.name, reverse=True)


def load_visitor_data(visitor_id):
    """加载来访者数据"""
    # 从 visitor_id 推导 case_id (V20260616001 -> C20260616001)
    case_id = visitor_id.replace('V', 'C', 1)
    case_file = DATA_DIR / f"{case_id}.json"

    visitor_data = {
        'visitor_id': visitor_id,
        'name': visitor_id,
        'phone': '',
        'counselor': '',
        'first_visit_date': '',
        'visit_count': 0,
        'last_visit_date': '',
        'crisis_status': '未评估',  # 大观评级状态
        'case_status': '未知',      # 案例状态
        'age': '',
        'gender': '',
        'occupation': ''
    }

    if case_file.exists():
        try:
            with open(case_file, 'r', encoding='utf-8') as f:
                case_data = json.load(f)

            # 基本信息
            basic_info = case_data.get('basic_info', {})
            visitor_data['name'] = basic_info.get('代号', visitor_id)
            visitor_data['age'] = basic_info.get('年龄', '')
            visitor_data['gender'] = basic_info.get('性别', '')
            visitor_data['occupation'] = basic_info.get('职业', '')
            visitor_data['phone'] = basic_info.get('联系方式', '')

            # 接访信息
            session_info = case_data.get('session_info', {})
            visit_date = session_info.get('接访日期', '')
            visitor_data['first_visit_date'] = visit_date
            visitor_data['last_visit_date'] = visit_date

            # 接访次数（从session_info中提取）
            visit_number = session_info.get('接访次数', '第1次')
            if '第' in visit_number and '次' in visit_number:
                try:
                    visitor_data['visit_count'] = int(visit_number.replace('第', '').replace('次', ''))
                except:
                    visitor_data['visit_count'] = 1

            # 咨询师
            visitor_data['counselor'] = session_info.get('咨询师', '')

            # 风险评估（大观评级）
            analyses = case_data.get('analyses', {})
            daguanpai = analyses.get('daguanpai', {})
            crisis_level = daguanpai.get('crisis_level', '')

            if crisis_level:
                crisis_map = {
                    'S': 'S级-高度危机',
                    'L': 'L级-中度危机',
                    'M': 'M级-低度危机',
                    'N': 'N级-无明显危机'
                }
                visitor_data['crisis_status'] = crisis_map.get(crisis_level, '未评估')
        except Exception as e:
            print(f"读取案例数据失败 {case_id}: {e}")

    # 读取来访者目录下的 visit_*.json 获取实际来访次数和案例状态
    visitor_folder = OUTPUT_DIR / visitor_id
    if visitor_folder.exists():
        visit_files = list(visitor_folder.glob('visit_*.json'))
        if visit_files:
            visitor_data['visit_count'] = len(visit_files)

            # 获取最后一次来访日期和案例状态
            latest_visit = None
            latest_visit_data = None

            for visit_file in visit_files:
                try:
                    with open(visit_file, 'r', encoding='utf-8') as f:
                        visit_data = json.load(f)
                        visit_date = visit_data.get('visit_date', '')
                        if visit_date:
                            if not latest_visit or visit_date > latest_visit:
                                latest_visit = visit_date
                                latest_visit_data = visit_data
                except:
                    pass

            if latest_visit:
                visitor_data['last_visit_date'] = latest_visit

            # 从最新来访记录获取案例状态
            if latest_visit_data:
                case_status = latest_visit_data.get('case_status', '')
                if case_status:
                    visitor_data['case_status'] = case_status
                else:
                    # 如果没有明确的case_status字段，默认设为"进行中"
                    visitor_data['case_status'] = '进行中'

    return visitor_data


@app.route('/api/visitors', methods=['GET'])
def get_visitors():
    """获取所有来访者列表"""
    try:
        visitor_folders = get_visitor_folders()
        visitors = []

        for folder in visitor_folders:
            visitor_id = folder.name
            visitor_data = load_visitor_data(visitor_id)
            visitors.append(visitor_data)

        return jsonify({
            'success': True,
            'visitors': visitors,
            'total': len(visitors)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """导出Excel文件"""
    try:
        visitor_folders = get_visitor_folders()
        visitors = []

        for folder in visitor_folders:
            visitor_id = folder.name
            visitor_data = load_visitor_data(visitor_id)
            visitors.append(visitor_data)

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "来访者列表"

        # 定义样式
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # 设置列标题
        headers = ['案例编号', '来访者', '性别', '年龄', '职业', '联系电话', '咨询师', '首访时间', '最近来访', '累计来访次数', '咨询状态', '案例状态']
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 设置列宽
        column_widths = [15, 18, 8, 15, 15, 15, 12, 12, 12, 12, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 设置行高
        ws.row_dimensions[1].height = 25

        # 填充数据
        for visitor in visitors:
            row_data = [
                visitor['visitor_id'],
                visitor['name'],
                visitor['gender'],
                visitor['age'],
                visitor['occupation'],
                visitor['phone'],
                visitor['counselor'],
                visitor['first_visit_date'],
                visitor['last_visit_date'],
                visitor['visit_count'],
                visitor['crisis_status'],
                visitor['case_status']
            ]
            ws.append(row_data)

            # 设置数据行样式
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.alignment = cell_alignment
                cell.border = border

            # 根据咨询状态（大观评级）设置颜色
            crisis_status = visitor['crisis_status']
            if 'S级' in crisis_status or '高度危机' in crisis_status:
                status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws.cell(row=row_num, column=11).fill = status_fill
            elif 'L级' in crisis_status or '中度危机' in crisis_status:
                status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                ws.cell(row=row_num, column=11).fill = status_fill
            elif 'M级' in crisis_status or '低度危机' in crisis_status:
                status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws.cell(row=row_num, column=11).fill = status_fill

            # 根据案例状态设置颜色
            case_status = visitor['case_status']
            if '进行中' in case_status:
                case_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws.cell(row=row_num, column=12).fill = case_fill
            elif '已结束' in case_status:
                case_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                ws.cell(row=row_num, column=12).fill = case_fill
            elif '暂停' in case_status:
                case_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                ws.cell(row=row_num, column=12).fill = case_fill

        # 保存到内存
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 生成文件名（包含导出时间）
        filename = f"来访者列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


if __name__ == '__main__':
    print("=" * 60)
    print("来访者列表数据服务启动")
    print(f"端口: 8770")
    print(f"输出目录: {OUTPUT_DIR}")
    print(f"数据目录: {DATA_DIR}")
    print("=" * 60)
    app.run(host='0.0.0.0', port=8770, debug=True)
