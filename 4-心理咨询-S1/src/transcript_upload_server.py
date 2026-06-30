#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
逐字稿上传服务器
端口: 8769
功能: 上传、下载、删除逐字稿文件
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import json
import os
import sys
import io
from pathlib import Path
from datetime import datetime
from werkzeug.utils import secure_filename
import subprocess
import shutil

# Windows GBK兼容性处理
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

app = Flask(__name__)
CORS(app)

PROJECT_ROOT = Path(__file__).parent.parent
VISITORS_DIR = PROJECT_ROOT / 'data' / 'visitors'

# 允许的逐字稿文件格式
ALLOWED_EXTENSIONS = {'txt', 'docx', 'xlsx', 'csv'}


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_file_size_str(size_bytes):
    """将字节转换为可读的文件大小"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.2f} MB"


def get_transcripts_dir(visitor_id, visit_id):
    """获取逐字稿文件存储目录"""
    transcripts_dir = VISITORS_DIR / visitor_id / 'visits' / visit_id / 'transcripts'
    transcripts_dir.mkdir(parents=True, exist_ok=True)
    return transcripts_dir


def get_visit_json_path(visitor_id, visit_id):
    """获取visit JSON文件路径"""
    return VISITORS_DIR / visitor_id / 'visits' / f'{visit_id}.json'


def parse_transcript_file(file_path, file_extension):
    """解析逐字稿文件，返回结构化数据"""
    transcript_data = []

    try:
        if file_extension == 'txt':
            # 解析TXT格式
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                lines = content.strip().split('\n')
                for line in lines:
                    if line.strip():
                        # 尝试识别发言人（格式：发言人：内容）
                        if '：' in line or ':' in line:
                            parts = line.split('：', 1) if '：' in line else line.split(':', 1)
                            if len(parts) == 2:
                                transcript_data.append({
                                    'speaker': parts[0].strip(),
                                    'content': parts[1].strip(),
                                    'timestamp': ''
                                })
                            else:
                                transcript_data.append({
                                    'speaker': '未知',
                                    'content': line.strip(),
                                    'timestamp': ''
                                })
                        else:
                            transcript_data.append({
                                'speaker': '未知',
                                'content': line.strip(),
                                'timestamp': ''
                            })

        elif file_extension == 'docx':
            # 解析DOCX格式
            from docx import Document
            doc = Document(file_path)
            for para in doc.paragraphs:
                text = para.text.strip()
                if text:
                    # 尝试识别发言人
                    if '：' in text or ':' in text:
                        parts = text.split('：', 1) if '：' in text else text.split(':', 1)
                        if len(parts) == 2:
                            transcript_data.append({
                                'speaker': parts[0].strip(),
                                'content': parts[1].strip(),
                                'timestamp': ''
                            })
                        else:
                            transcript_data.append({
                                'speaker': '未知',
                                'content': text,
                                'timestamp': ''
                            })
                    else:
                        transcript_data.append({
                            'speaker': '未知',
                            'content': text,
                            'timestamp': ''
                        })

        elif file_extension in ['xlsx', 'csv']:
            # 解析Excel/CSV格式
            if file_extension == 'xlsx':
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    ws = wb.active

                    headers = [cell.value for cell in ws[1]]
                    speaker_col = content_col = time_col = None

                    for i, header in enumerate(headers):
                        if header and ('发言人' in str(header) or 'speaker' in str(header).lower()):
                            speaker_col = i
                        elif header and ('内容' in str(header) or 'content' in str(header).lower()):
                            content_col = i
                        elif header and ('时间' in str(header) or 'timestamp' in str(header).lower()):
                            time_col = i

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # 跳过空行
                        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                            continue

                        # 安全访问列，避免索引越界
                        speaker = '未知'
                        content = ''
                        timestamp = ''

                        if speaker_col is not None and speaker_col < len(row) and row[speaker_col]:
                            speaker = str(row[speaker_col])
                        if content_col is not None and content_col < len(row) and row[content_col]:
                            content = str(row[content_col])
                        if time_col is not None and time_col < len(row) and row[time_col]:
                            timestamp = str(row[time_col])

                        if content.strip():
                            transcript_data.append({
                                'speaker': speaker,
                                'content': content,
                                'timestamp': timestamp
                            })
                except ImportError:
                    raise Exception("需要安装openpyxl库: pip install openpyxl")
            else:  # CSV
                import csv
                with open(file_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        speaker = row.get('发言人', row.get('speaker', '未知'))
                        content = row.get('内容', row.get('content', ''))
                        timestamp = row.get('时间', row.get('timestamp', ''))

                        if content.strip():
                            transcript_data.append({
                                'speaker': speaker,
                                'content': content,
                                'timestamp': timestamp
                            })

    except Exception as e:
        print(f"解析文件失败: {e}")
        raise

    return transcript_data


def trigger_html_generation():
    """触发HTML生成"""
    try:
        creation_flags = 0
        if sys.platform == 'win32':
            creation_flags = subprocess.CREATE_NO_WINDOW

        script_path = PROJECT_ROOT / 'src' / 'generate_visit_details.py'
        subprocess.Popen(
            [sys.executable, str(script_path)],
            cwd=str(PROJECT_ROOT),
            creationflags=creation_flags,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        print("  [触发] HTML生成任务已提交")
    except Exception as e:
        print(f"  [WARN] 触发HTML生成失败: {e}")


@app.route('/upload', methods=['POST'])
def upload_transcript():
    """上传逐字稿文件"""
    try:
        visitor_id = request.form.get('visitor_id')
        visit_id = request.form.get('visit_id')

        print(f"\n[上传] 逐字稿上传请求")
        print(f"  visitor_id: {visitor_id}")
        print(f"  visit_id: {visit_id}")

        if not visitor_id or not visit_id:
            return jsonify({'success': False, 'error': '缺少visitor_id或visit_id'}), 400

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400

        file = request.files['file']
        if file.filename == '' or not allowed_file(file.filename):
            return jsonify({'success': False, 'error': '文件格式不支持'}), 400

        # 保存原始文件
        original_filename = file.filename  # 保留原始文件名（包括中文）

        # 直接从原始文件名提取扩展名
        if '.' in original_filename:
            file_extension = original_filename.rsplit('.', 1)[1].lower()
        else:
            file_extension = 'txt'  # 默认扩展名

        transcripts_dir = get_transcripts_dir(visitor_id, visit_id)

        # 生成唯一文件名
        transcript_id = f"trans_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        saved_filename = f"{transcript_id}.{file_extension}"
        file_path = transcripts_dir / saved_filename

        file.save(str(file_path))
        file_size = file_path.stat().st_size

        print(f"  ✓ 文件已保存: {saved_filename} ({get_file_size_str(file_size)})")

        # 解析逐字稿内容
        try:
            transcript_content = parse_transcript_file(str(file_path), file_extension)
            print(f"  ✓ 解析成功: {len(transcript_content)} 条记录")
        except Exception as e:
            print(f"  ✗ 解析失败: {e}")
            transcript_content = []

        # 更新visit JSON
        visit_json_path = get_visit_json_path(visitor_id, visit_id)
        if not visit_json_path.exists():
            return jsonify({'success': False, 'error': f'找不到来访记录: {visit_id}'}), 404

        with open(visit_json_path, 'r', encoding='utf-8') as f:
            visit_data = json.load(f)

        # 确保transcript_files字段存在
        if 'transcript_files' not in visit_data['case_data']:
            visit_data['case_data']['transcript_files'] = []

        # 添加文件记录
        file_record = {
            'transcript_id': transcript_id,
            'filename': original_filename,  # 使用原始文件名
            'saved_filename': saved_filename,
            'file_size': file_size,
            'file_size_str': get_file_size_str(file_size),
            'file_format': file_extension.upper(),
            'upload_time': datetime.now().isoformat(),
            'content': transcript_content,
            'description': request.form.get('description', f'{original_filename}')  # 使用原始文件名
        }

        visit_data['case_data']['transcript_files'].append(file_record)
        visit_data['metadata']['updated_at'] = datetime.now().isoformat()

        with open(visit_json_path, 'w', encoding='utf-8') as f:
            json.dump(visit_data, f, ensure_ascii=False, indent=2)

        print(f"  ✓ 逐字稿记录已保存")

        # 触发HTML重新生成
        trigger_html_generation()

        return jsonify({
            'success': True,
            'message': '逐字稿上传成功',
            'transcript_id': transcript_id,
            'content_count': len(transcript_content)
        })

    except Exception as e:
        print(f"✗ 上传失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/download/<visitor_id>/<visit_id>/<transcript_id>', methods=['GET'])
def download_transcript(visitor_id, visit_id, transcript_id):
    """下载逐字稿文件"""
    try:
        # 读取visit JSON获取文件信息
        visit_json_path = get_visit_json_path(visitor_id, visit_id)
        if not visit_json_path.exists():
            return jsonify({'error': '找不到来访记录'}), 404

        with open(visit_json_path, 'r', encoding='utf-8') as f:
            visit_data = json.load(f)

        # 查找文件记录
        file_record = None
        for record in visit_data['case_data'].get('transcript_files', []):
            if record['transcript_id'] == transcript_id:
                file_record = record
                break

        if not file_record:
            return jsonify({'error': '找不到逐字稿文件'}), 404

        # 获取文件路径
        transcripts_dir = get_transcripts_dir(visitor_id, visit_id)
        file_path = transcripts_dir / file_record['saved_filename']

        if not file_path.exists():
            return jsonify({'error': '文件不存在'}), 404

        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=file_record['filename']
        )

    except Exception as e:
        print(f"✗ 下载失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/delete', methods=['POST'])
def delete_transcript():
    """删除逐字稿文件"""
    try:
        data = request.json
        visitor_id = data.get('visitor_id')
        visit_id = data.get('visit_id')
        transcript_id = data.get('transcript_id')

        print(f"\n[删除] 逐字稿删除请求")
        print(f"  visitor_id: {visitor_id}")
        print(f"  visit_id: {visit_id}")
        print(f"  transcript_id: {transcript_id}")

        if not all([visitor_id, visit_id, transcript_id]):
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400

        # 读取visit JSON
        visit_json_path = get_visit_json_path(visitor_id, visit_id)
        if not visit_json_path.exists():
            return jsonify({'success': False, 'error': '找不到来访记录'}), 404

        with open(visit_json_path, 'r', encoding='utf-8') as f:
            visit_data = json.load(f)

        # 查找并删除文件记录
        transcript_files = visit_data['case_data'].get('transcript_files', [])
        file_record = None
        new_files = []

        for record in transcript_files:
            if record['transcript_id'] == transcript_id:
                file_record = record
            else:
                new_files.append(record)

        if not file_record:
            return jsonify({'success': False, 'error': '找不到逐字稿记录'}), 404

        # 删除物理文件
        transcripts_dir = get_transcripts_dir(visitor_id, visit_id)
        file_path = transcripts_dir / file_record['saved_filename']
        if file_path.exists():
            os.unlink(file_path)
            print(f"  ✓ 文件已删除: {file_record['saved_filename']}")

        # 更新JSON
        visit_data['case_data']['transcript_files'] = new_files
        visit_data['metadata']['updated_at'] = datetime.now().isoformat()

        with open(visit_json_path, 'w', encoding='utf-8') as f:
            json.dump(visit_data, f, ensure_ascii=False, indent=2)

        print(f"  ✓ 记录已删除")

        # 触发HTML重新生成
        trigger_html_generation()

        return jsonify({'success': True, 'message': '逐字稿删除成功'})

    except Exception as e:
        print(f"✗ 删除失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    print("=" * 60)
    print("逐字稿上传服务器")
    print("=" * 60)
    print(f"端口: 8769")
    print(f"数据目录: {VISITORS_DIR}")
    print(f"支持格式: TXT, DOCX, XLSX, CSV")
    print(f"API端点:")
    print(f"  - POST /upload                            上传逐字稿文件")
    print(f"  - GET  /download/<vid>/<visid>/<tid>      下载逐字稿文件")
    print(f"  - POST /delete                             删除逐字稿文件")
    print(f"  - GET  /health                             健康检查")
    print("=" * 60)
    app.run(host='0.0.0.0', port=8769, debug=False)
