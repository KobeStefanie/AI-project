#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""T恤工厂财务核算模板生成器"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "d:/AI-项目/12-成本控制"
os.makedirs(BASE + "/工资管理", exist_ok=True)
os.makedirs(BASE + "/成本核算", exist_ok=True)

def BD():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def H(c, v, bg="2E75B6", merge=None, ws=None):
    c.value = v
    c.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BD()
    if merge and ws:
        ws.merge_cells(c.coordinate + ":" + merge)

def N(c, v="", bg="FFFFFF", bold=False, al="left", fc="000000"):
    c.value = v
    c.font = Font(name="微软雅黑", size=10, bold=bold, color=fc)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=al, vertical="center")
    c.border = BD()

def Y(ws, r, col, v=""):  # 黄色输入格
    c = ws.cell(r, col, value=v)
    c.fill = PatternFill("solid", fgColor="FFFFC0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BD()

def G(ws, r, col, formula, fmt="#,##0.00"):  # 绿色公式格
    c = ws.cell(r, col, value=formula)
    c.fill = PatternFill("solid", fgColor="E2EFDA")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = BD()
    c.number_format = fmt

def tax(r):  # 个税公式（月度简算，f-string转义Excel数组花括号）
    return (f"=ROUND(MAX((MAX(0,K{r}-L{r}-M{r}-N{r}-5000))"
            f"*{{0.03,0.1,0.2,0.25,0.3,0.35,0.45}}"
            f"-{{0,210,1410,2660,4410,7160,15160}},0),2)")

# ═══ 工资核算模板 ═══════════════════════════════════════════
def build_payroll():
    wb = openpyxl.Workbook()

    # Sheet1: 参数设置
    ws = wb.active; ws.title = "参数设置"
    for col, w in zip("ABCD", [26, 14, 6, 28]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30
    H(ws.cell(1,1), "工资核算参数（黄色格可改）", bg="1F3864", merge="D1", ws=ws)
    rows = [
        (3,"基本参数",None,None,None,"1F3864"),
        (4,"公司名称","新疆某T恤有限公司","","修改为实际名称",None),
        (5,"发薪日",25,"日","每月几号发",None),
        (6,"月标准工作天数",26,"天","用于日薪计算",None),
        (7,"最低工资(元/月)",2400,"元","按新疆标准填",None),
        (8,"加班系数",None,None,None,"1F3864"),
        (9,"平日延时",1.5,"倍","超8小时",None),
        (10,"休息日",2.0,"倍","周六周日",None),
        (11,"节假日",3.0,"倍","法定假日",None),
        (12,"社保比例（个人）",None,None,None,"1F3864"),
        (13,"养老","8%","","",None),
        (14,"医疗","2%","","",None),
        (15,"失业","0.5%","","",None),
        (16,"公积金","5%","","新疆部分地区可能不缴",None),
        (17,"社保比例（单位）",None,None,None,"1F3864"),
        (18,"养老","16%","","",None),
        (19,"医疗","8%","","含生育",None),
        (20,"失业","0.5%","","",None),
        (21,"工伤","0.5%","","按行业费率",None),
        (22,"公积金","5%","","",None),
    ]
    for rd in rows:
        r,lbl,val,unit,note,sec = rd
        ws.row_dimensions[r].height = 20
        if sec:
            H(ws.cell(r,1), lbl, bg=sec, merge=f"D{r}", ws=ws)
        else:
            N(ws.cell(r,1), lbl, bold=True)
            Y(ws, r, 2, val); N(ws.cell(r,3), unit, al="center")
            N(ws.cell(r,4), note, fc="595959")

    # Sheet2: 员工档案
    ws2 = wb.create_sheet("员工档案")
    emp_hdrs = [("工号",8),("姓名",10),("部门",12),("岗位",12),
                ("工资类型",11),("底薪(元)",11),("计件单价",12),
                ("社保基数",12),("银行卡号",22),("开户行",14)]
    ws2.row_dimensions[1].height = 30
    for j,(name,w) in enumerate(emp_hdrs,1):
        ws2.column_dimensions[chr(64+j)].width = w
        H(ws2.cell(1,j), name)
    demo = [("001","张三","裁剪","裁床工","计件",2400,0.8,2400,"6225881234567890","建设银行"),
            ("002","李四","缝纫","缝纫工","计件",2400,1.2,2400,"6217001234567890","工商银行"),
            ("003","王五","质检","质检员","计时",4500,0,4500,"6228481234567890","农业银行"),
            ("004","赵六","仓储","仓管员","综合",3200,0,3200,"6236681234567890","中国银行")]
    for i,row in enumerate(demo,2):
        ws2.row_dimensions[i].height = 20
        for j,v in enumerate(row,1): Y(ws2, i, j, v)
    N(ws2.cell(6,1),"... 向下继续填写，共500行",fc="888888",bg="F5F5F5")
    ws2.merge_cells("A6:J6")
    ws2.freeze_panes = "A2"

    # Sheet3: 月度工资核算
    ws3 = wb.create_sheet("月度工资核算")
    ph = [("工号",8),("姓名",10),("部门",12),("出勤天数",10),("底薪应发",12),
          ("计件数量",10),("计件工资",12),("加班小时",10),("加班工资",12),
          ("绩效奖金",12),("应发合计",12),("养老个人",11),("医疗个人",11),
          ("公积金个人",12),("个人所得税",12),("实  发",13)]
    ws3.row_dimensions[1].height = 42
    for j,(name,w) in enumerate(ph,1):
        ws3.column_dimensions[chr(64+j)].width = w
        H(ws3.cell(1,j), name)
    for i in range(2,6):
        ws3.row_dimensions[i].height = 22
        # 工号姓名部门 从员工档案读
        G(ws3,i,1,f"=员工档案!A{i}",fmt="@")
        G(ws3,i,2,f"=员工档案!B{i}",fmt="@")
        G(ws3,i,3,f"=员工档案!C{i}",fmt="@")
        Y(ws3,i,4,26)                                  # 出勤天数
        G(ws3,i,5,f"=员工档案!F{i}/参数设置!$B$6*D{i}")  # 底薪
        Y(ws3,i,6)                                     # 计件数量
        G(ws3,i,7,f"=F{i}*员工档案!G{i}")              # 计件工资
        Y(ws3,i,8,0)                                   # 加班小时
        G(ws3,i,9,f"=H{i}*(员工档案!F{i}/参数设置!$B$6/参数设置!$B$7)*参数设置!$B$9")
        Y(ws3,i,10,0)                                  # 绩效奖金
        G(ws3,i,11,f"=E{i}+G{i}+I{i}+J{i}")           # 应发合计
        G(ws3,i,12,f"=员工档案!H{i}*0.08")             # 养老
        G(ws3,i,13,f"=员工档案!H{i}*0.02")             # 医疗
        G(ws3,i,14,f"=员工档案!H{i}*0.05")             # 公积金
        G(ws3,i,15,tax(i))                             # 个税
        G(ws3,i,16,f"=K{i}-L{i}-M{i}-N{i}-O{i}")      # 实发
    # 合计行
    rr = 6; ws3.row_dimensions[rr].height = 26
    H(ws3.cell(rr,1),"合  计",bg="1F3864",merge=f"C{rr}",ws=ws3)
    for j in range(4,17):
        if j in (4,6,8,10): N(ws3.cell(rr,j),"",bg="DDDDDD")
        else: G(ws3,rr,j,f"=SUM({chr(64+j)}2:{chr(64+j)}{rr-1})")
    ws3.freeze_panes = "D2"
    # 添加说明
    ws3.cell(8,1).value="🟡 黄色列=手工录入  🟢 绿色列=自动计算  向下复制行即可扩展至500行"
    ws3.cell(8,1).font = Font(name="微软雅黑",size=10,color="595959")
    ws3.merge_cells("A8:P8")

    wb.save(BASE + "/工资管理/工资核算模板.xlsx")
    print("  OK: 工资核算模板.xlsx")


# ═══ 标准成本核算模板 ══════════════════════════════════════
def build_cost():
    wb = openpyxl.Workbook()

    # Sheet1: 产品成本目标
    ws = wb.active; ws.title = "产品成本目标"
    for col,w in zip("ABCDEFG",[8,24,12,12,12,14,18]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32
    H(ws.cell(1,1),"T恤产品标准成本目标一览",bg="1F3864",merge="G1",ws=ws)
    for j,h in enumerate(["代码","产品名称","售价(元)","目标毛利率","目标成本","实际成本","状态"],1):
        H(ws.cell(2,j), h)
    prods = [("A001","基础款白色圆领",45,"35%"),("A002","基础款彩色圆领",52,"35%"),
             ("B001","加重款(180g+)",68,"38%"),("B002","Polo衫",85,"40%"),
             ("C001","定制印花款",95,"42%")]
    for i,(code,name,price,margin) in enumerate(prods,3):
        ws.row_dimensions[i].height = 22
        N(ws.cell(i,1),code,al="center"); N(ws.cell(i,2),name)
        Y(ws,i,3,price); Y(ws,i,4,margin)
        N(ws.cell(i,5),"（见成本卡）",fc="888888")
        Y(ws,i,6,"")
        G(ws,i,7,f'=IF(F{i}="","待录入",IF(F{i}<=C{i}*(1-0.35),"✅达标","❌超支"))',fmt="@")

    # Sheet2: 标准成本卡
    ws2 = wb.create_sheet("标准成本卡(A001)")
    for col,w in zip("ABCDE",[12,26,16,14,14]):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[1].height = 32
    H(ws2.cell(1,1),"单品标准成本卡 — A001基础款白色圆领T恤",bg="1F3864",merge="E1",ws=ws2)
    for j,h in enumerate(["成本类别","项目明细","标准用量","单价/费率","标准成本(元)"],1):
        H(ws2.cell(2,j), h)
    items = [
        ("直接材料","面料(180g棉布)","200g/件","30元/kg",6.0),
        ("直接材料","辅料(线+标签)","—","—",1.5),
        ("直接材料","包装(袋+吊牌)","—","—",0.8),
        ("直接人工","裁剪(计件)","3分钟/件","0.5元/分钟",1.5),
        ("直接人工","缝纫(计件)","8分钟/件","0.5元/分钟",4.0),
        ("直接人工","整烫包装","2分钟/件","0.5元/分钟",1.0),
        ("制造费用","设备折旧分摊","—","—",1.2),
        ("制造费用","水电费分摊","—","—",0.5),
        ("制造费用","工厂管理费","—","—",0.8),
    ]
    prev=""
    for i,(cat,item,qty,price2,cost) in enumerate(items,3):
        ws2.row_dimensions[i].height = 22
        N(ws2.cell(i,1),cat if cat!=prev else "",bold=(cat!=prev),al="center")
        N(ws2.cell(i,2),item); Y(ws2,i,3,qty); Y(ws2,i,4,price2); Y(ws2,i,5,cost)
        prev=cat
    rr=len(items)+3; ws2.row_dimensions[rr].height = 26
    H(ws2.cell(rr,1),"标准总成本",bg="1F3864",merge=f"D{rr}",ws=ws2)
    G(ws2,rr,5,f"=SUM(E3:E{rr-1})")

    # Sheet3: 月度差异分析
    ws3 = wb.create_sheet("月度差异分析")
    for col,w in zip("ABCDEFGHI",[8,22,12,14,14,14,14,12,22]):
        ws3.column_dimensions[col].width = w
    ws3.row_dimensions[1].height = 32
    H(ws3.cell(1,1),"月度成本差异分析（录入实际成本后自动计算）",bg="1F3864",merge="I1",ws=ws3)
    for j,h in enumerate(["代码","产品名称","产量(件)","标准单位成本","标准总成本","实际总成本","差异(元)","差异率","结论"],1):
        H(ws3.cell(2,j), h)
    for i,(code,name) in enumerate([("A001","基础款白色"),("A002","基础款彩色"),("B001","加重款")],3):
        ws3.row_dimensions[i].height = 22
        N(ws3.cell(i,1),code,al="center"); N(ws3.cell(i,2),name)
        Y(ws3,i,3,""); Y(ws3,i,4,""); Y(ws3,i,6,"")
        G(ws3,i,5,f"=IF(C{i}*D{i}=0,\"\",C{i}*D{i})")
        G(ws3,i,7,f"=IF(OR(E{i}=\"\",F{i}=\"\"),\"\",F{i}-E{i})")
        G(ws3,i,8,f"=IF(E{i}=0,\"\",ROUND(G{i}/E{i}*100,1))",fmt="0.0")
        cond = (f"=IF(H{i}=\"\",\"待录入\","
                f"IF(H{i}>5,\"❌超支需分析\","
                f"IF(H{i}>0,\"⚠️轻微超支\","
                f"IF(H{i}>-5,\"✅正常\",\"💰成本节约\"))))")
        G(ws3,i,9,cond,fmt="@")

    wb.save(BASE + "/成本核算/标准成本核算表.xlsx")
    print("  OK: 标准成本核算表.xlsx")


if __name__ == "__main__":
    print("生成财务核算模板（纯Excel，离线可用）...")
    build_payroll()
    build_cost()
    print(f"\n📁 {BASE}/工资管理/工资核算模板.xlsx")
    print("  => {BASE}/成本核算/标准成本核算表.xlsx")
    print("\n✅ 这两个文件是标准xlsx格式，无需网络，无需Claude，Excel直接打开即用")

